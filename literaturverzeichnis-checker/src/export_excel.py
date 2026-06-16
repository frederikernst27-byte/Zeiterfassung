"""Schreibt die Prüfergebnisse als Excel-Tabelle."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .classify import STATUS_MINOR_ISSUES, STATUS_NOT_FOUND, STATUS_OK, STATUS_UNCLEAR, Result

STATUS_COLORS = {
    STATUS_OK: "C6EFCE",
    STATUS_MINOR_ISSUES: "FFEB9C",
    STATUS_NOT_FOUND: "FFC7CE",
    STATUS_UNCLEAR: "D9D9D9",
}

HEADERS = [
    "Nr.", "Original-Zitat", "Status", "Gefundene Quelle",
    "Abweichungen", "Prüfmethode", "Konfidenz (%)",
]


def export_to_excel(results: list[Result], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Literaturpruefung"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        ws.append([
            r.number,
            r.original_citation,
            r.status,
            r.found_source or "",
            "; ".join(r.discrepancies),
            r.method,
            round(r.confidence, 1),
        ])
        fill_color = STATUS_COLORS.get(r.status)
        if fill_color:
            ws.cell(row=ws.max_row, column=3).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    widths = [6, 50, 28, 45, 45, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(output_path)
